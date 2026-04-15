import subprocess
import time
import os
import shutil
import httpx  
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# --- CONFIGURATION PATHS ---
GDB_PATH = r"C:\NXP\S32DS_ARM_v2.2\S32DS\build_tools\gcc_v6.3\gcc-6.3-arm32-eabi\bin\arm-none-eabi-gdb.exe"
PESERVER_PATH = r"C:\NXP\S32DS_ARM_v2.2\eclipse\plugins\com.pemicro.debug.gdbjtag.pne_6.0.9.202509241532\win32\pegdbserver_console.exe"
DEVICE_NAME = "NXP_S32K1xx_S32K146F1M0M11"
EXCEL_PATH = r"F:\OneDrive - Raptee Energy Pvt Ltd\MC ELECTRICAL CHECKLIST.xlsx"

FMLITE_DIR = r"C:\NXP\FreeMASTER 3.2\FreeMASTER Lite"
FMLITE_EXE = os.path.join(FMLITE_DIR, "fmlite.exe")
ACTIVE_CONFIG = os.path.join(FMLITE_DIR, "config.json")
NEWEOL_TEMPLATE = os.path.join(FMLITE_DIR, "config_new_eol.json")
BIKE_TEMPLATE = os.path.join(FMLITE_DIR, "config_bike.json")

class LogRequest(BaseModel):
    unique_id: str
    trace_id: str
    tester: str
    wave: str
    fan: str
    hv: str
    offset: str
    throttle_status: str
    mode: str = "offline"

class FlashRequest(BaseModel):
    action: str
    file_path: str = None

def kill_process(name):
    subprocess.run(["taskkill", "/IM", name, "/F"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

def manage_fmlite():
    print(">> Cleaning up FreeMASTER Lite processes...")
    kill_process("fmlite.exe")
    time.sleep(3) 
    if os.path.exists(ACTIVE_CONFIG):
        print(">> Launching FreeMASTER Lite with active config...")
        subprocess.Popen([FMLITE_EXE, "config.json"], cwd=FMLITE_DIR, creationflags=0x08000000)
    else:
        print(">> ERROR: config.json missing")

def swap_config(target):
    source = NEWEOL_TEMPLATE if target == "new_eol" else BIKE_TEMPLATE
    if os.path.exists(source):
        print(f">> Swapping config to: {target}")
        shutil.copy(source, ACTIVE_CONFIG)

@app.post("/flash")
async def trigger_flash(request: FlashRequest):
    target_elf = request.file_path if request.file_path else r"F:/MC_FLASH_FILES/NewEoL.elf"
    file_name = os.path.basename(target_elf)
    kill_process("fmlite.exe")
    kill_process("pegdbserver_console.exe")

    with open("flash_mcu.gdb", "w") as f:
        f.write(f"target extended-remote localhost:7224\n"
                f"file \"{target_elf.replace('\\', '/')}\"\n"
                f"load\n"
                f"monitor reset\n"
                f"quit\n")

    pes = subprocess.Popen(
        [PESERVER_PATH, "-startserver", f"-device={DEVICE_NAME}", "-port", "7224", "-F", "0"],
        stdout=subprocess.PIPE, text=True
    )

    start = time.time()
    while time.time() - start < 15:
        line = pes.stdout.readline()
        if "All Servers Running" in line:
            break

    result = subprocess.run([GDB_PATH, "-x", "flash_mcu.gdb"], capture_output=True, text=True)
    pes.terminate()

    if result.returncode == 0:
        swap_config("new_eol" if "NewEoL" in file_name else "bike")
        manage_fmlite()
        return {"status": "success", "file": file_name}
    
    return {"status": "error", "message": result.stderr}

@app.post("/log_excel")
async def log_excel(data: LogRequest):
    try:
        timestamp = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        
        # --- GLOBAL PASS CONDITION ---
        tests_to_check = [data.wave, data.fan, data.hv, data.throttle_status]
        overall = "PASS" if all(status == "PASS" for status in tests_to_check) else "FAIL"

        if not os.path.exists(EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.append(["S.NO.", "UNIQUE QR ID", "MCU - ID NO.", "DATE", "TEST 1 WAVE FORMS", "TEST 2 FAN CHECK", "TEST 3 HV TEST", "MC OFFSET VALUES", "TEST 4 CAN CHECK", "TESTED BY", "Test Passed for Electrical"])
        else:
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active

        # Check for existing UID to update or append
        target_row = None
        for row in range(2, ws.max_row + 1):
            excel_uid = str(ws.cell(row=row, column=2).value).strip()
            if excel_uid == str(data.unique_id).strip():
                target_row = row
                break

        row_content = [
            target_row - 1 if target_row else ws.max_row,
            data.unique_id, data.trace_id, timestamp,
            data.wave, data.fan, data.hv, data.offset,
            data.throttle_status, data.tester, overall
        ]

        if target_row:
            for col, value in enumerate(row_content, start=1):
                ws.cell(row=target_row, column=col, value=value)
        else:
            ws.append(row_content)

        wb.save(EXCEL_PATH)
        return {"status": "success", "message": "Logged to Excel successfully"}
    except Exception as e:
        return {"status": "error", "message": f"Excel Error: {str(e)}"}

@app.post("/log_gallus")
async def log_gallus(data: LogRequest):
    try:
        # --- GLOBAL PASS CONDITION ---
        tests_to_check = [data.wave, data.fan, data.hv, data.throttle_status]
        overall = "PASS" if all(status == "PASS" for status in tests_to_check) else "FAIL"

        # Zoho Configuration
        base_url = "https://www.zohoapis.in/creator/custom/raptee.hv/Insert_MC_Logs?publickey=hTQHa5Eb2DErtYpEujzhs43Su"
        headers = {"content-type": "application/json"}
        
        params = {
            "pUnique_QR_ID": data.unique_id,
            "pTest_No_1": data.trace_id,
            "pTEST_1_50_DUTY_CYCLE_WAVE_FORMS": data.wave,
            "pTEST_2_MC_FAN_CHECK": data.fan,
            "pTEST_3_HV_TEST": data.hv,
            "pMC_OFFSET_VALUES": data.offset,
            "pTEST_4_CAN_CHECK": data.throttle_status,
            "pTest_Passed_for_Electrical": overall, # Set based on condition above
            "pTEST_3_FLASH": "PASS",
            "pCONDITION_PASS_FAIL": overall,        # Set based on condition above
            "pTESTED_BY": data.tester
        }

        async with httpx.AsyncClient() as client:
            response = await client.post(base_url, json=params, headers=headers, timeout=15.0)
            
            if response.status_code == 200:
                return {"status": "success", "message": "Uploaded to Gallus (Zoho) successfully"}
            else:
                return {"status": "error", "message": f"Zoho API Error: {response.status_code}"}
                
    except Exception as e:
        return {"status": "error", "message": f"Cloud Sync Failed: {str(e)}"}

if __name__ == "__main__":
    import uvicorn
    # Initial cleanup and launch of FreeMASTER Lite
    manage_fmlite()
    uvicorn.run(app, host="127.0.0.1", port=5000)